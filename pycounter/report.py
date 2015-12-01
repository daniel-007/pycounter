"""COUNTER journal and book reports and associated functions"""

from __future__ import absolute_import

import datetime
import logging
import re

import arrow
import six

from pycounter import csvhelper
from pycounter.constants import METRICS, CODES, REPORT_DESCRIPTIONS
from pycounter.exceptions import UnknownReportTypeError, PycounterException
from pycounter.helpers import convert_covered, convert_date_run, \
    convert_date_column, last_day, next_month, format_stat


class CounterReport(object):
    """
    a COUNTER usage statistics report.

    Iterate over the report object to get its rows (each of which is a
    :class:`CounterBook <CounterBook>` or :class:`CounterJournal
    <CounterJournal>` instance.

    :param metric: metric being tracked by this report. For database
        reports (which have multiple metrics per report), this should be
        set to `None`.

    :param report_type: type of report (e.g., "JR1", "BR2")

    :param report_version: COUNTER version

    :param customer: name of customer on report

    :param institutional_identifier: unique ID assigned by vendor for
        customer

    :param period: tuple of datetime.date objects corresponding to the
        beginning and end of the covered range

    :param date_run: date the COUNTER report was generated

    """

    def __init__(self, report_type=None, report_version=4, metric=None,
                 customer=None, institutional_identifier=None,
                 period=(None, None), date_run=None):
        self.pubs = []
        self.report_type = report_type
        self.report_version = report_version
        self.metric = metric
        self.customer = customer
        self.institutional_identifier = institutional_identifier
        self.period = period
        if date_run is None:
            self.date_run = datetime.date.today()
        else:
            self.date_run = date_run
        self.year = None

    def __str__(self):
        return (
            "CounterReport %s version %s for date range %s to %s" %
            (self.report_type,
             self.report_version,
             self.period[0], self.period[1])
        )

    def __iter__(self):
        return iter(self.pubs)

    def write_tsv(self, path):
        """
        Output report to a COUNTER 4 TSV file.

        :param path: location to write file
        """
        lines = self.as_generic()
        with csvhelper.UnicodeWriter(path, delimiter='\t') as writer:
            writer.writerows(lines)

    def as_generic(self):
        """
        Output report as list of lists, containing cells that would appear
        in COUNTER report (suitable for writing as CSV, TSV, etc.)
        """
        output_lines = []
        rep_type = ""
        for name, code in CODES.items():
            if code == self.report_type[0:2]:
                rep_type = name

        report_name = ("%s Report %s (R%s)" %
                       (rep_type, self.report_type[-1], self.report_version))
        output_lines.append([report_name,
                             REPORT_DESCRIPTIONS[self.report_type]])
        output_lines.append([self.customer])
        output_lines.append([self.institutional_identifier])
        output_lines.append([u'Period covered by Report:'])
        period = "%s to %s" % (
            self.period[0].strftime('%Y-%m-%d'),
            self.period[1].strftime('%Y-%m-%d')
        )
        output_lines.append([period])
        output_lines.append([u'Date run:'])
        output_lines.append([self.date_run.strftime('%Y-%m-%d')])
        output_lines.append(self._table_header())
        output_lines.append(self._totals_line())

        for pub in self.pubs:
            output_lines.append(pub.as_generic())

        return output_lines

    def _totals_line(self):
        """
        Generate Totals line for COUNTER report, as list of cells
        """
        # FIXME: don't hardcode JR1 values
        total_cells = [
            u'Total for all journals',
            u'',
        ]
        platforms = set(resource.platform for resource in self.pubs)
        if len(platforms) == 1:
            total_cells.append(platforms.pop())
        else:
            total_cells.append(u'')
        total_cells.extend([u''] * 4)
        total_usage = 0
        pdf_usage = 0
        html_usage = 0

        number_of_months = len(
            arrow.Arrow.range('month',
                              arrow.Arrow.fromdate(self.period[0]),
                              arrow.Arrow.fromdate(self.period[1])))
        month_data = [0] * number_of_months
        for pub in self.pubs:
            pdf_usage += pub.pdf_total
            html_usage += pub.html_total
            for month, data in enumerate(pub):
                total_usage += data[2]
                month_data[month] += data[2]
        total_cells.append(six.text_type(total_usage))
        total_cells.append(six.text_type(html_usage))
        total_cells.append(six.text_type(pdf_usage))
        total_cells.extend(six.text_type(d) for d in month_data)

        return total_cells

    def _table_header(self):
        """
        Generate header for COUNTER table for this report, as list of cells
        """
        # FIXME: don't hardcode JR1 values.
        header_cells = [
            u'Journal',
            u'Publisher',
            u'Platform',
            u'Journal DOI',
            u'Proprietary Identifier',
            u'Print ISSN',
            u'Online ISSN',
            u'Reporting Period Total',
            u'Reporting Period HTML',
            u'Reporting Period PDF',
        ]
        for d_obj in arrow.Arrow.range('month',
                                       arrow.Arrow.fromdate(self.period[0]),
                                       arrow.Arrow.fromdate(self.period[1])):
            header_cells.append(d_obj.strftime('%b-%Y'))

        return header_cells


class CounterEresource(six.Iterator):
    # pylint: disable=too-few-public-methods
    """
    base class for COUNTER statistics lines

    Iterating returns (first_day_of_month, metric, usage) tuples.

    :param period: two-tuple of datetime.date objects corresponding
        to the beginning and end dates of the covered range

    :param metric: metric tracked by this report. Should be a value
        from pycounter.report.METRICS dict.

    :param month_data: a list containing usage data for this
        resource, as (datetime.date, usage) tuples

    :param title: title of the resource

    :param publisher: name of the resource's publisher

    :param platform: name of the platform providing the resource

    """

    def __init__(self, period=None, metric=None, month_data=None,
                 title="", platform="", publisher=""):
        self.period = period

        self.metric = metric
        self._monthdata = []
        self._full_data = []
        if month_data is not None:
            for item in month_data:
                self._full_data.append(item)

        if title:
            self.title = title
        if platform:
            self.platform = platform
        if publisher:
            self.publisher = publisher

    def __iter__(self):
        if self._full_data:
            for item in self._full_data:
                yield (item[0], self.metric, item[1])
        else:
            current_month = self.period[0]
            month_data_iter = iter(self._monthdata)
            while current_month < self.period[1]:
                current_usage = next(month_data_iter)
                yield (current_month, self.metric, current_usage)
                current_month = next_month(current_month)


class CounterJournal(CounterEresource):
    """
    statistics for a single electronic journal.

    :param period: two-tuple of datetime.date objects corresponding
        to the beginning and end dates of the covered range

    :param metric: the metric tracked by this statistics line.
        (Should probably always be "FT Article Requests" for
        CounterJournal objects, as long as only JR1 is supported.)

    :param issn: eJournal's print ISSN

    :param eissn: eJournal's eISSN

    :param month_data: a list containing usage data for this
        journal, as (datetime.date, usage) tuples

    :param title: title of the resource

    :param publisher: name of the resource's publisher

    :param platform: name of the platform providing the resource

    :param html_total: total HTML usage for this title for reporting period

    :param pdf_total: total PDF usage for this title for reporting period

    """

    def __init__(self, period=None, metric=METRICS[u"JR1"],
                 issn=None, eissn=None, month_data=None,
                 title="", platform="", publisher="", html_total=0,
                 pdf_total=0, doi="", proprietary_id=""):
        super(CounterJournal, self).__init__(period, metric, month_data,
                                             title, platform, publisher)
        self.html_total = html_total
        self.pdf_total = pdf_total
        self.doi = doi
        self.proprietary_id = proprietary_id

        if issn is not None:
            self.issn = issn
        else:
            self.issn = ''


        if eissn is not None:
            self.eissn = eissn
        else:
            self.eissn = ''

    def __str__(self):
        return """<CounterJournal %s, publisher %s,
        platform %s>""" % (self.title, self.publisher, self.platform)

    def as_generic(self):
        """
        return data for this line as list of COUNTER report cells
        """
        data_line = [
            self.title,
            self.publisher,
            self.platform,
            self.doi,
            self.proprietary_id,
            self.issn,
            self.eissn,
        ]
        total_usage = 0
        month_data = []
        for data in self:
            total_usage += data[2]
            month_data.append(six.text_type(data[2]))
        data_line.append(six.text_type(total_usage))
        data_line.append(six.text_type(self.html_total))
        data_line.append(six.text_type(self.pdf_total))
        data_line.extend(month_data)
        return data_line


class CounterBook(CounterEresource):
    # pylint: disable=too-few-public-methods
    """
    statistics for a single electronic book.

    :ivar isbn: eBook's ISBN

    :ivar issn: eBook's ISSN (if any)

    :param month_data: a list containing usage data for this
        book, as (datetime.date, usage) tuples

    :param title: title of the resource

    :param publisher: name of the resource's publisher

    :param platform: name of the platform providing the resource

    """

    def __init__(self, period=None, metric=None, month_data=None,
                 title="", platform="", publisher="", isbn=None, issn=None):
        super(CounterBook, self).__init__(period, metric, month_data,
                                          title, platform, publisher)
        self.eissn = None

        if isbn is not None:
            self.isbn = isbn

        if issn is not None:
            self.issn = issn

    def __str__(self):
        return """<CounterBook %s (ISBN: %s), publisher %s,
        platform %s>""" % (self.title, self.isbn, self.publisher,
                           self.platform)


class CounterDatabase(CounterEresource):
    # pylint: disable=too-few-public-methods
    """a COUNTER database report line"""

    def __init__(self, period=None, metric=None, month_data=None,
                 title="", platform="", publisher=""):
        super(CounterDatabase, self).__init__(period, metric, month_data,
                                              title, platform, publisher)


def parse(filename, filetype=None):
    """Parse a COUNTER file, first attempting to determine type

    Returns a :class:`CounterReport <CounterReport>` object.

    :param filename: path to COUNTER report to load and parse.
    :param filetype: type of file provided, one of "csv", "tsv", "xlsx".
        If set to None (the default), an attempt will be made to
        detect the correct type, first from the file extension, then from
        the file's contents.

    """
    if filetype is None:
        if filename.endswith('.tsv'):
            filetype = 'tsv'
        elif filename.endswith('.xlsx'):
            filetype = 'xlsx'
        elif filename.endswith('.csv'):
            filetype = 'csv'
        else:
            with open(filename, 'rb') as file_obj:
                first_bytes = file_obj.read(2)
                if first_bytes == b"PK":
                    filetype = 'xlsx'
                else:
                    content = file_obj.read()
                    if b'\t' in content:
                        filetype = 'tsv'
                    else:
                        filetype = 'csv'

    if filetype == 'tsv':
        return parse_separated(filename, '\t')
    elif filetype == 'xlsx':
        return parse_xlsx(filename)
    elif filetype == 'csv':
        return parse_separated(filename, ',')
    else:
        raise PycounterException("Unknown file type %s" % filetype)


def parse_xlsx(filename):
    """Parse a COUNTER file in Excel format.

    Invoked automatically by ``parse``.

    :param filename: path to XLSX-format COUNTER report file.

    """
    from openpyxl import load_workbook
    with open(filename, 'rb') as xlsx_file:
        workbook = load_workbook(xlsx_file)
        worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
        row_it = worksheet.iter_rows()
        split_row_list = ([cell.value if cell.value is not None else ""
                           for cell in row] for row in row_it)

    return parse_generic(split_row_list)


def parse_separated(filename, delimiter):
    """Open COUNTER CSV/TSV report with given filename and delimiter
    and parse into a CounterReport object

    Invoked automatically by :py:func:`parse`.

    :param filename: path to delimited COUNTER report file.

    :param delimiter: character (such as ',' or '\\\\t') used as the
        delimiter for this file

    """
    with csvhelper.UnicodeReader(filename,
                                 delimiter=delimiter) as report_reader:
        return parse_generic(report_reader)


def parse_generic(report_reader):
    """Takes an iterator of COUNTER report rows and
    returns a CounterReport object

    :param report_reader: a iterable object that yields lists COUNTER
        data formatted as tabular lists

    """
    report = CounterReport()

    line1 = six.next(report_reader)

    rt_match = re.match(
        r'.*(Journal|Book|Database) Report (\d(?: GOA)?) ?\(R(\d)\)',
        line1[0])
    if rt_match:
        report.report_type = (CODES[rt_match.group(1)] +
                              rt_match.group(2))
        report.report_version = int(rt_match.group(3))

    # noinspection PyTypeChecker
    report.metric = METRICS.get(report.report_type)

    report.customer = six.next(report_reader)[0]

    if report.report_version == 4:
        inst_id_line = six.next(report_reader)
        if inst_id_line:
            report.institutional_identifier = inst_id_line[0]

        six.next(report_reader)

        covered_line = six.next(report_reader)
        report.period = convert_covered(covered_line[0])

    six.next(report_reader)

    date_run_line = six.next(report_reader)
    report.date_run = convert_date_run(date_run_line[0])

    header = six.next(report_reader)
    first_date_col = 10 if report.report_version == 4 else 5
    if report.report_type in ('BR1', 'BR2') and report.report_version == 4:
        first_date_col = 8
    elif report.report_type == 'DB1' and report.report_version == 4:
        first_date_col = 6
    elif report.report_type == 'DB2' and report.report_version == 4:
        first_date_col = 5
    year = int(header[first_date_col].split('-')[1])
    if year < 100:
        year += 2000

    report.year = year

    if report.report_version == 4:
        countable_header = header[0:8]
        for col in header[8:]:
            if col:
                countable_header.append(col)
        last_col = len(countable_header)
    else:
        last_col = 0
        for val in header:
            if 'YTD' in val:
                break
            last_col += 1

        start_date = datetime.date(year, 1, 1)
        end_date = last_day(convert_date_column(header[last_col - 1]))
        report.period = (start_date, end_date)

    if report.report_type != 'DB1':
        six.next(report_reader)

    if report.report_type == 'DB2':
        six.next(report_reader)

    for line in report_reader:
        issn = None
        eissn = None
        html_total = 0
        pdf_total = 0
        doi = ""
        prop_id = ""
        if not line:
            continue
        if report.report_version == 4:
            if report.report_type.startswith('JR1'):
                old_line = line
                line = line[0:3] + line[5:7] + line[10:last_col]
                doi = old_line[3]
                prop_id = old_line[4]
                html_total = int(old_line[8])
                pdf_total = int(old_line[9])
                issn = line[3].strip()
                eissn = line[4].strip()

            elif report.report_type in ('BR1', 'BR2'):
                line = line[0:3] + line[5:7] + line[8:last_col]
            elif report.report_type in ('DB1', 'DB2'):
                # format coincidentally works for these. This is a kludge
                # so leaving this explicit...
                pass
        else:
            if report.report_type.startswith('JR1'):
                html_total = int(line[-2])
                pdf_total = int(line[-1])
                issn = line[3].strip()
                eissn = line[4].strip()
            line = line[0:last_col]

        logging.debug(line)
        title = line[0]
        publisher = line[1]
        platform = line[2]
        month_data = []
        curr_month = report.period[0]
        for data in line[5:]:
            month_data.append((curr_month, format_stat(data)))
            curr_month = next_month(curr_month)
        if report.report_type:
            if report.report_type.startswith('JR'):
                report.pubs.append(CounterJournal(report.period,
                                                  report.metric,
                                                  month_data=month_data,
                                                  title=title,
                                                  publisher=publisher,
                                                  platform=platform,
                                                  doi=doi,
                                                  issn=issn,
                                                  eissn=eissn,
                                                  proprietary_id=prop_id,
                                                  html_total=html_total,
                                                  pdf_total=pdf_total))
            elif report.report_type.startswith('BR'):
                report.pubs.append(CounterBook(report.period,
                                               report.metric,
                                               month_data=month_data,
                                               title=title,
                                               publisher=publisher,
                                               platform=platform))
            elif report.report_type.startswith('DB'):
                report.pubs.append(CounterDatabase(report.period,
                                                   line[3],
                                                   month_data=month_data,
                                                   title=title,
                                                   publisher=publisher,
                                                   platform=platform))
            else:
                raise UnknownReportTypeError(report.report_type)

    return report
